from __future__ import annotations
import io, re, json, tempfile
from pathlib import Path
from typing import Dict, Tuple
from pdf2image import convert_from_bytes
from paddleocr import PaddleOCR
import numpy as np
import cv2
import pandas as pd
from openpyxl import load_workbook

def load_config(path: str = "config.json") -> dict:
    import json
    with open(path, encoding="utf-8") as f:
        return json.load(f)

_OCR = None
def _get_ocr(lang: str = "japan", use_angle_cls: bool = True, 
             det_db_thresh: float = 0.2, det_db_box_thresh: float = 0.3,
             rec_batch_num: int = 6) -> PaddleOCR:
    """
    PaddleOCR初期化（シングルトンパターン・高精度設定）
    
    Args:
        lang: OCR言語 (japan=日本語, en=英語, ch=中国語など)
        use_angle_cls: 文字の回転検出を有効化 (True=回転文字も認識/精度UP, False=高速)
        det_db_thresh: テキスト検出閾値（低いほど小さい文字も検出、0.2=最高感度）
        det_db_box_thresh: ボックス信頼度閾値（低いほど広範囲検出、0.3=広範囲）
        rec_batch_num: 認識バッチサイズ（小さいほど精度UP、6=精度重視）
    
    Returns:
        PaddleOCRインスタンス
    
    ※精度最大化設定:
    - use_angle_cls=True: 回転文字認識
    - det_db_thresh=0.2: 最高検出感度
    - det_db_box_thresh=0.3: 広範囲検出
    - rec_batch_num=6: 精度重視バッチ
    - use_space_char=True: スペース文字認識
    - drop_score=0.3: 低信頼度文字も採用
    """
    global _OCR
    if _OCR is None:
        _OCR = PaddleOCR(
            lang=lang,
            use_angle_cls=use_angle_cls,
            det_db_thresh=det_db_thresh,
            det_db_box_thresh=det_db_box_thresh,
            rec_batch_num=rec_batch_num,
            use_space_char=True,
            drop_score=0.3,
            show_log=False
        )
    return _OCR

def process_pdf_bytes(pdf_bytes: bytes, cfg: dict) -> Tuple[Dict, str]:
    """
    PDFファイルからOCRでテキストを抽出し、指定フィールドを取得
    
    Args:
        pdf_bytes: PDFファイルのバイナリデータ
        cfg: config.jsonの設定辞書
    
    Returns:
        (抽出フィールド辞書, 全テキスト)
    
    ※OCR精度向上のための設定 (config.json):
    - ocr.dpi: 200→300-400 (解像度UP=精度UP、処理時間増)
    - performance.max_width: 1000→1200-1500 (画像幅UP=精度UP、メモリ消費増)
    - paddle.use_angle_cls: false→true (回転文字認識=精度UP)
    - performance.max_pages_per_pdf: 処理するページ数制限 (0=全ページ)
    """
    # config.jsonから設定を取得（バランス型高精度設定）
    dpi = cfg.get("ocr", {}).get("dpi", 300)  # OCR解像度（300=高精度バランス型）
    max_pages = cfg.get("performance", {}).get("max_pages_per_pdf", 3)  # 処理ページ数制限（3ページ）
    max_width = cfg.get("performance", {}).get("max_width", 1500)  # 画像最大幅（1500=高精度）
    use_angle_cls = cfg.get("paddle", {}).get("use_angle_cls", True)  # 回転検出（精度UP）
    lang = cfg.get("paddle", {}).get("lang", "japan")  # OCR言語
    det_db_thresh = cfg.get("paddle", {}).get("det_db_thresh", 0.3)  # 検出閾値（バランス型）
    det_db_box_thresh = cfg.get("paddle", {}).get("det_db_box_thresh", 0.5)  # ボックス閾値（標準）
    rec_batch_num = cfg.get("paddle", {}).get("rec_batch_num", 10)  # バッチサイズ（バランス）

    images = convert_from_bytes(pdf_bytes, dpi=dpi)
    texts = []
    fields = {}
    for i, img in enumerate(images):
        if i >= max_pages:
            break
        # 画像をNumPy配列に変換
        img_np = np.array(img)
        h, w = img_np.shape[:2]
        
        # 画像幅を制限してメモリ消費を抑制（精度とのバランス）
        if w > max_width:
            scale = max_width / w
            new_w = int(w * scale)
            new_h = int(h * scale)
            img_np = cv2.resize(img_np, (new_w, new_h), interpolation=cv2.INTER_AREA)
        
        # グレースケール変換
        if len(img_np.shape) == 3:
            gray = cv2.cvtColor(img_np, cv2.COLOR_BGR2GRAY)
        else:
            gray = img_np.copy()
        
        # シンプル高精度前処理: 最小限の処理で最大の精度
        # ガウシアンブラーでノイズ除去（バイラテラルより安定）
        gray = cv2.GaussianBlur(gray, (3, 3), 0)
        
        # 2値化なし: PaddleOCRはグレースケールのままの方が精度が高い
        img_np = gray
        
        # OCR実行（バランス型パラメータ）
        ocr = _get_ocr(lang, use_angle_cls, det_db_thresh, det_db_box_thresh, rec_batch_num)
        res = ocr.ocr(img_np, cls=use_angle_cls)
        page_text = "\n".join([t[1][0] for line in res for t in (line if isinstance(line, list) else [line])])
        texts.append(page_text)
        # 簡易フィールド抽出例（法人名・契約電力）
        for k in cfg.get("targets", {}):
            for kw in cfg.get("excel_input", {}).get("label_keywords", {}).get(k, []):
                match = re.search(rf"{kw}[:：]?\s*([^\s\n]+)", page_text)
                if match:
                    fields[k] = match.group(1)
    return fields, "\n\n".join(texts)

def process_excel_bytes(excel_bytes: bytes, cfg: dict) -> Tuple[Dict, pd.DataFrame]:
    from io import BytesIO
    excel_io = BytesIO(excel_bytes)
    wb = load_workbook(excel_io, data_only=True)
    sheet_name = cfg.get("excel_cell_map", {}).get("sheet", wb.sheetnames[0])
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    data = ws.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)
    fields = {}
    # 簡易フィールド抽出例（法人名・契約電力）
    for k in cfg.get("targets", {}):
        for kw in cfg.get("excel_input", {}).get("label_keywords", {}).get(k, []):
            for col in df.columns:
                if kw in str(col):
                    fields[k] = df[col].iloc[0]
    return fields, df

def write_to_excel(list_of_fields: list, cfg: dict) -> str:
    from openpyxl import load_workbook
    template_path = Path("template_output.xlsx")
    if not template_path.exists():
        return ""
    wb = load_workbook(template_path)
    sheet_name = cfg.get("excel_cell_map", {}).get("sheet", wb.sheetnames[0])
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # 1行目から順に代入（例：B2, G2 → B3, G3 → ...）
    start_row = 2  # 1行目はヘッダー想定
    for idx, fields in enumerate(list_of_fields):
        for key, cell in cfg.get("excel_cell_map", {}).items():
            # 列記号と行番号を分離
            import re
            m = re.match(r"([A-Z]+)(\d+)", cell)
            if m:
                col, base_row = m.group(1), int(m.group(2))
                target_cell = f"{col}{start_row + idx}"
                if key in fields:
                    ws[target_cell] = fields[key]
    out_path = "output_combined.xlsx"
    wb.save(out_path)
    return out_path
